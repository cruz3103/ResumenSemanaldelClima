import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

# Cargar archivo
df = pd.read_excel('datos_meteorologicos.xlsx')

# Función para limpiar columnas
def limpiar_columna(col, unidad):
    return pd.to_numeric(
        df[col].astype(str)
        .str.replace(unidad, '', regex=False)
        .str.replace(',', '.')
        .str.strip(),
        errors='coerce'
    )

# Procesamiento de datos
df['Fecha'] = pd.to_datetime(df['Fecha'])

# Calcular semana del ciclo y número de ciclo
df['Semana del ciclo'] = np.nan
df['Ciclo'] = np.nan

# Detectar ciclos por año
ciclos = []
for year in sorted(df['Fecha'].dt.year.unique()):
    base = pd.to_datetime(f'{year}-07-18')
    if base.weekday() != 0:
        inicio_ciclo = base + pd.DateOffset(days=(7 - base.weekday()))
    else:
        inicio_ciclo = base
    fin_ciclo = inicio_ciclo + pd.DateOffset(weeks=52)
    ciclos.append((inicio_ciclo, fin_ciclo))

# Asignar semana del ciclo y número de ciclo
for i, (inicio, fin) in enumerate(ciclos):
    mask = (df['Fecha'] >= inicio) & (df['Fecha'] < fin)
    df.loc[mask, 'Semana del ciclo'] = ((df.loc[mask, 'Fecha'] - inicio).dt.days // 7) + 1
    df.loc[mask, 'Ciclo'] = i + 1

df['Semana del ciclo'] = df['Semana del ciclo'].astype('Int64')
df['Ciclo'] = df['Ciclo'].astype('Int64')

# Limpieza de variables
df['Temperature'] = limpiar_columna('Temperature (°C)', '°C')
df['Humidity'] = limpiar_columna('Humidity (%)', '%')
df['Precip'] = limpiar_columna('Precip. Accum. (mm)', 'mm')
df['Solar'] = limpiar_columna('Solar Radiation (W/m²)', 'w/m²')
df['Wind'] = limpiar_columna('Speed (km/h)', 'km/h')
df['Precip_Estimada'] = limpiar_columna('Precip. Rate. (mm/hr)', 'mm') * 0.25
df['Radiacion_Estimada'] = limpiar_columna('Solar Radiation (W/m²)', 'W/m²') * 0.25

# Calcular HDD como (T - T_base) * 0.25
T_base = 10
df['HDD_15min'] = (df['Temperature'] - T_base) * 0.25

# Asegurarse de que 'Fecha' esté ordenada
df = df.sort_values('Fecha')

# Calcular HDD diario
df['Dia'] = pd.to_datetime(df['Fecha'].dt.date)
hdd_diario = df.groupby('Dia')['HDD_15min'].sum().reset_index()
hdd_diario = hdd_diario.sort_values('Dia')
hdd_diario['HDD_60d_acumulado'] = hdd_diario['HDD_15min'].rolling(window=60, min_periods=1).sum()

# Fusionar con el DataFrame principal
df = df.merge(hdd_diario[['Dia', 'HDD_60d_acumulado']], on='Dia', how='left')

# --- Precipitación estimada acumulada últimos 60 días ---
precip_diaria = df.groupby('Dia')['Precip_Estimada'].sum().reset_index()
precip_diaria = precip_diaria.sort_values('Dia')
precip_diaria['Precip_60d_acumulada'] = precip_diaria['Precip_Estimada'].rolling(window=60, min_periods=1).sum()

# --- Radiación estimada acumulada últimos 60 días ---
rad_diaria = df.groupby('Dia')['Radiacion_Estimada'].sum().reset_index()
rad_diaria = rad_diaria.sort_values('Dia')
rad_diaria['Radiacion_60d_acumulada'] = rad_diaria['Radiacion_Estimada'].rolling(window=60, min_periods=1).sum()

# Fusionar al DataFrame principal
df = df.merge(precip_diaria[['Dia', 'Precip_60d_acumulada']], on='Dia', how='left')
df = df.merge(rad_diaria[['Dia', 'Radiacion_60d_acumulada']], on='Dia', how='left')

# Semana
df['Semana'] = df['Fecha'].dt.to_period('W').apply(lambda r: r.start_time)

# Días con sol
dias_con_sol = df.groupby('Dia')['Solar'].max().reset_index()
dias_con_sol['Soleado'] = dias_con_sol['Solar'] > 0
dias_con_sol['Semana'] = pd.to_datetime(dias_con_sol['Dia']).dt.to_period('W').apply(lambda r: r.start_time)
conteo_dias_soleados = dias_con_sol.groupby('Semana')['Soleado'].sum().reset_index()
conteo_dias_soleados.columns = ['Semana', 'Días soleados (Solar > 0 W/m²)']

# Días con heladas
temperaturas_dia = df.groupby(['Dia', 'Semana'])['Temperature'].min().reset_index()
temperaturas_dia['Dia_frio'] = temperaturas_dia['Temperature'] < 4
conteo_dias_frios = temperaturas_dia.groupby('Semana')['Dia_frio'].sum().reset_index()
conteo_dias_frios.columns = ['Semana', 'Días con heladas (< 4 °C)']

# Resumen semanal
resumen = df.groupby('Semana').agg(
    Ciclo=('Ciclo', 'first'),
    temperatura_media=('Temperature', 'mean'),
    temperatura_min=('Temperature', 'min'),
    temperatura_max=('Temperature', 'max'),
    humedad_media=('Humidity', 'mean'),
    viento_promedio=('Wind', 'mean'),
    precipitacion_total=('Precip', 'max'),
    solar_promedio=('Solar', 'mean'),
    precipitacion_estimada_total=('Precip_Estimada', 'sum'),
    radiacion_estimada_total=('Radiacion_Estimada', 'sum'),
    hdd_total=('HDD_15min', 'sum'),
    precipitacion_60d_final=('Precip_60d_acumulada', 'last'),
    radiacion_60d_final=('Radiacion_60d_acumulada', 'last')
).reset_index()

# Acumulado de HDD en las últimas 8 semanas (~60 días)
resumen = resumen.sort_values('Semana')
resumen['HDD acumulado últimos 60 días'] = resumen['hdd_total'].rolling(window=8, min_periods=1).sum()

# Agregar conteo de heladas y días soleados
resumen = resumen.merge(conteo_dias_frios, on='Semana', how='left')
resumen = resumen.merge(conteo_dias_soleados, on='Semana', how='left')

# Renombrar columnas
resumen = resumen.rename(columns={
    'temperatura_media': 'Temperatura media (°C)',
    'temperatura_min': 'Temperatura mínima (°C)',
    'temperatura_max': 'Temperatura máxima (°C)',
    'humedad_media': 'Humedad media (%)',
    'viento_promedio': 'Viento promedio (km/h)',
    'precipitacion_total': 'Precipitación máx. diaria (mm)',
    'solar_promedio': 'Radiación solar promedio (W/m²)',
    'precipitacion_estimada_total': 'Precipitación estimada (mm)',
    'radiacion_estimada_total': 'Radiación estimada (Wh/m²)',
    'hdd_total': 'Grados-día calefacción (HDD)',
    'precipitacion_60d_final': 'Precipitación estimada acumulada últimos 60 días (mm)',
    'radiacion_60d_final': 'Radiación estimada acumulada últimos 60 días (Wh/m²)',
})

# Reordenar columnas según preferencia del usuario
cols = resumen.columns.tolist()

# Seleccionar y reordenar

orden_personalizado = [
    'Semana',
    'Ciclo',
    'Temperatura media (°C)',
    'Temperatura mínima (°C)',
    'Temperatura máxima (°C)',
    'Humedad media (%)',
    'Viento promedio (km/h)',
    'Precipitación máx. diaria (mm)',
    'Radiación solar promedio (W/m²)',
    'Precipitación estimada (mm)',
    'Precipitación estimada acumulada últimos 60 días (mm)',
    'Radiación estimada (Wh/m²)',
    'Radiación estimada acumulada últimos 60 días (Wh/m²)',
    'Grados-día calefacción (HDD)',
    'HDD acumulado últimos 60 días',
    'Días con heladas (< 4 °C)',
    'Días soleados (Solar > 0 W/m²)'
]

# Reordenar solo si están todas las columnas
resumen = resumen[[col for col in orden_personalizado if col in resumen.columns]]


##print(resumen[['Semana', 'HDD acumulado últimos 60 días']].head(140))

# Descripción de columnas
descripcion = pd.DataFrame({
    'Columna': resumen.columns,
    'Descripción': [
        'Inicio de la semana (lunes)',
        'Número de ciclo agrícola (inicia desde 1, luego 2, etc.)',
        'Promedio semanal de temperatura en grados Celsius',
        'Temperatura mínima semanal en °C',
        'Temperatura máxima semanal en °C',
        'Promedio semanal de humedad relativa en porcentaje',
        'Promedio semanal de velocidad del viento en km/h',
        'Total de precipitación semanal acumulada en mm',
        'Promedio semanal de radiación solar en W/m²',
        'Total semanal estimado de precipitación (mm), calculado como tasa × 0.25',
        'Acumulado de precipitación estimada en los 60 días anteriores',
        'Total semanal estimado de energía solar recibida en Wh/m²',
        'Acumulado de radiación estimada en los 60 días anteriores',
        'Total semanal de Heating Degree DaysGrados-día de calefacción: suma de (T_base - T) * 0.25 cada 15 minutos',
        'Acumulado de grados-día de calefacción en los 60 días anteriores',
        'Cantidad de días con heladas (temperatura menor a 4 °C)',
        'Cantidad de días con radiación solar registrada (mayor a 0 W/m²)'
    ]
})

# Gráficos
plt.figure(figsize=(12, 6))
x = np.arange(len(resumen['Semana']))
bar_width = 0.25
plt.bar(x - bar_width, resumen['Temperatura media (°C)'], width=bar_width, label='Temp. media (°C)')
plt.bar(x, resumen['Humedad media (%)'], width=bar_width, label='Humedad media (%)')
plt.bar(x + bar_width, resumen['Viento promedio (km/h)'], width=bar_width, label='Viento promedio (km/h)')
plt.title('Resumen semanal: temperatura, humedad y viento')
plt.xlabel('Semana')
plt.ylabel('Valor')
plt.xticks(x, resumen['Semana'].dt.strftime('%Y-%m-%d'), rotation=45)
plt.legend()
plt.grid(axis='y')
plt.tight_layout()
plt.savefig('grafico_clima_barras.png')
plt.close()

plt.figure(figsize=(12, 6))
plt.bar(resumen['Semana'].dt.strftime('%Y-%m-%d'), resumen['Radiación solar promedio (W/m²)'], color='orange')
plt.title('Radiación solar promedio semanal')
plt.xlabel('Semana')
plt.ylabel('W/m²')
plt.xticks(rotation=45)
plt.grid(axis='y')
plt.tight_layout()
plt.savefig('grafico_solar_barras.png')
plt.close()

plt.figure(figsize=(12, 6))
plt.bar(resumen['Semana'].dt.strftime('%Y-%m-%d'), resumen['Días con heladas (< 4 °C)'], color='skyblue')
plt.title('Días con heladas por semana (temperatura < 4 °C)')
plt.xlabel('Semana')
plt.ylabel('Número de días')
plt.xticks(rotation=45)
plt.grid(axis='y')
plt.tight_layout()
plt.savefig('grafico_heladas_barras.png')
plt.close()

# --- GRÁFICAS DE ACUMULADOS 60 DÍAS ---

# Gráfico 1: Precipitación acumulada últimos 60 días
plt.figure(figsize=(12, 6))
plt.plot(resumen['Semana'], resumen['Precipitación estimada acumulada últimos 60 días (mm)'], marker='o', color='blue')
plt.title('Precipitación estimada acumulada (últimos 60 días)')
plt.xlabel('Semana')
plt.ylabel('mm')
plt.grid(True)
plt.xticks(rotation=45)
plt.tight_layout()
plt.savefig('grafico_precip_60d.png')
plt.close()

# Gráfico 2: Radiación acumulada últimos 60 días
plt.figure(figsize=(12, 6))
plt.plot(resumen['Semana'], resumen['Radiación estimada acumulada últimos 60 días (Wh/m²)'], marker='o', color='orange')
plt.title('Radiación acumulada estimada (últimos 60 días)')
plt.xlabel('Semana')
plt.ylabel('Wh/m²')
plt.grid(True)
plt.xticks(rotation=45)
plt.tight_layout()
plt.savefig('grafico_radiacion_60d.png')
plt.close()

# Gráfico 3: HDD acumulado últimos 60 días
plt.figure(figsize=(12, 6))
plt.plot(resumen['Semana'], resumen['HDD acumulado últimos 60 días'], marker='o', color='green')
plt.title('Grados-día calefacción acumulados (últimos 60 días)')
plt.xlabel('Semana')
plt.ylabel('HDD')
plt.grid(True)
plt.xticks(rotation=45)
plt.tight_layout()
plt.savefig('grafico_hdd_60d.png')
plt.close()

fig, ax1 = plt.subplots(figsize=(14, 7))

# Eje principal: Precipitación
color1 = 'blue'
ax1.set_xlabel('Semana')
ax1.set_ylabel('Precipitación (mm)', color=color1)
ax1.plot(resumen['Semana'], resumen['Precipitación estimada acumulada últimos 60 días (mm)'],
         marker='o', color=color1, label='Precipitación')
ax1.tick_params(axis='y', labelcolor=color1)

# Segundo eje: Radiación
ax2 = ax1.twinx()
color2 = 'orange'
ax2.set_ylabel('Radiación (Wh/m²)', color=color2)
ax2.plot(resumen['Semana'], resumen['Radiación estimada acumulada últimos 60 días (Wh/m²)'],
         marker='s', color=color2, label='Radiación')
ax2.tick_params(axis='y', labelcolor=color2)

# Tercer eje: HDD
ax3 = ax1.twinx()
color3 = 'green'
ax3.spines.right.set_position(("axes", 1.12))  # desplaza el eje
ax3.set_ylabel('HDD', color=color3)
ax3.plot(resumen['Semana'], resumen['HDD acumulado últimos 60 días'],
         marker='^', color=color3, label='HDD')
ax3.tick_params(axis='y', labelcolor=color3)

# Usar fechas en el eje X correctamente
ax1.xaxis.set_tick_params(rotation=45)
fig.autofmt_xdate()  # ajusta las etiquetas de fecha automáticamente

# Título y leyenda
plt.title('Comparativa: Acumulados 60 días - Precipitación, Radiación y HDD')

# Leyenda combinada
lines, labels = [], []
for ax in [ax1, ax2, ax3]:
    line, label = ax.get_legend_handles_labels()
    lines += line
    labels += label
ax1.legend(lines, labels, loc='upper left')

plt.tight_layout()
plt.savefig('grafico_comparado_60d.png')
plt.close()


# Guardar en Excel
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

try:
    with pd.ExcelWriter('resumen_semanal_clima.xlsx', engine='openpyxl') as writer:
        resumen.to_excel(writer, index=False, sheet_name='Resumen semanal')
        descripcion.to_excel(writer, index=False, sheet_name='Descripción')

        writer.book.create_sheet('Gráficos')
        sheet = writer.book['Gráficos']
        img1 = Image('grafico_clima_barras.png')
        img2 = Image('grafico_solar_barras.png')
        img3 = Image('grafico_heladas_barras.png')
        sheet.add_image(img1, 'A1')
        sheet.add_image(img2, 'A30')
        sheet.add_image(img3, 'A59')

         # Nueva hoja con acumulados 60 días
        writer.book.create_sheet('Acumulados 60d')
        sheet_acum = writer.book['Acumulados 60d']
        img4 = Image('grafico_precip_60d.png')
        img5 = Image('grafico_radiacion_60d.png')
        img6 = Image('grafico_hdd_60d.png')
        sheet_acum.add_image(img4, 'A1')
        sheet_acum.add_image(img5, 'A30')
        sheet_acum.add_image(img6, 'A59')

        # nueva hoja con las comparacion de las tres graficas
        # Nueva hoja con gráfico comparativo acumulado 60 días
        writer.book.create_sheet('Comparativa 60d')
        sheet_comp = writer.book['Comparativa 60d']
        img7 = Image('grafico_comparado_60d.png')
        sheet_comp.add_image(img7, 'A1')   

        writer.book.save('resumen_semanal_clima.xlsx')

    print("✅ Archivo actualizado con gráficos y resumen semanal.")

except PermissionError:
    print("❌ No se pudo guardar el archivo. Asegúrate de cerrar 'resumen_semanal_clima.xlsx'.")
